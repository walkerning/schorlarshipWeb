# -*- coding: utf-8 -*-
from __future__ import print_function

import re
import os
import sys
import getpass
import logging
import argparse
from collections import namedtuple

import chardet
import requests
from openpyxl import load_workbook, Workbook
from pyquery import PyQuery as pq

logger = logging.getLogger("export_info")
logging.basicConfig(format="[%(levelname)s] %(name)s: %(message)s")

# URLs for login
login_url = "https://info.tsinghua.edu.cn:443/Login" # info login
manage_tab_url = "http://info.tsinghua.edu.cn/tag.e28e0eefe5c8efa2.render.userLayoutRootNode.uP?uP_sparam=focusedTabID&focusedTabID=9&uP_sparam=mode&mode=view&_meta_focusedId=9" # 管理界面
dept_manage_url = "http://jxxxfw.cic.tsinghua.edu.cn/roam.do3?type=srch&id=307" # 院系管理列表
ee_manage_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/index.jsp?xsh=023" # 电子系院系管理
audit_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/check/admin.jsp" # 院系审核

# URLs
student_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/check/check.jsp"
xf_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/fill/xf_add.jsp"
yf_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/fill/yf_add.jsp"
yg_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/fill/yg_add.jsp"
scho_delete_url = "http://zhjw.cic.tsinghua.edu.cn/bksjzd/xsgzz/zxj/fill/{type}_delete.jsp?xh={student_id}&dm={dm}"

user_agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36"
info_url = "info.tsinghua.edu.cn"

def login(username, password, sess):
    data = {
        "userName": username,
        "password": password
    }
    headers = {
        "User-Agent": user_agent,
        "Referer": info_url
    }
    sess.post(login_url, data=data, headers=headers, allow_redirects=False)

class Exporter(object):
    def __init__(self, username, password, backup_dir=None):
        self.s = None
        s = requests.Session()
        login(username, password, s)
        headers = {
            "Host": info_url,
            "User-Agent": user_agent,
            "Upgrade-Insecure-Requests": "1"
        }
        manage_res = s.get(manage_tab_url, headers=headers)
        # get the ticket
        found = re.search(r"jxxxfw.cic.tsinghua.edu.cn/admin/zhcx.jsp\?rootid=944\&amp;ticket=([a-zA-Z0-9]+)", manage_res.text)
        assert found, "Login error, check the password."
        ticket = found.group(1)
        # logger.debug("Ticket: %s", ticket)
        headers = {
            "Referer": manage_tab_url,
            "User-Agent": user_agent,
        }
        res0 = s.get("http://jxxxfw.cic.tsinghua.edu.cn/admin/zhcx.jsp?rootid=944&ticket={}".format(ticket), headers=headers)
        # Here: 6 cookies got, info, meta, jxxxfw
        u_content0 = res0.content.decode(chardet.detect(res0.content)["encoding"])
        if u_content0.find(u"您的登陆已失效") >= 0:
            raise AssertionError("jxxxfw.cic/admin/zhcx.jsp 登陆错误")
        s.get(dept_manage_url, headers=headers) # zhjw cookie
        s.get(ee_manage_url, headers=headers)

        s.get(audit_url, headers=headers)

        # REDO: I don't know why it need redo... the first time fail. u_content1 will be `用户信息不存在或类型不正确`
        s.get(dept_manage_url, headers=headers) # zhjw cookie
        # u_content1 = res1.content.decode(chardet.detect(res1.content)["encoding"])
        s.get(ee_manage_url, headers=headers)
        # u_content2 = res2.content.decode(chardet.detect(res2.content)["encoding"])

        s.get(audit_url, headers=headers)
        # TODO: multiple page
        # res_content = res.content.decode(chardet.detect(res.content)["encoding"])
        self.s = s

        self.backup_dir = os.path.abspath(os.path.expanduser(backup_dir)) if backup_dir else None
        if self.backup_dir:
            print(("备份将被存至目录: {}. 请注意会覆盖原有该目录下的同名文件. "
                   "如果该目录下已有同名文件, 请先将该目录做备份, 再继续.").format(self.backup_dir))
            raw_input("按任意键继续...")
            if not os.path.isdir(self.backup_dir):
                logger.info("创建备份目录: {}".format(self.backup_dir))
                os.mkdir(self.backup_dir)

    def get_student_list(self):
        res = self.s.get(audit_url, params={"curPage": 1})
        res_content = res.content
        doc = pq(res_content)
        final_page_match = re.search('<a href=["\']admin.jsp\?curPage=([^"\']+)["\']>*?<img src=["\']/bksjzd/pic/end1.gif["\']', res_content).group(1)
        logger.info("一共 {} 页记录".format(final_page_match))
        lst = []
        for page in range(1, int(final_page_match) + 1):
            res = self.s.get(audit_url, params={"curPage": str(page)})
            doc = pq(res.content)
            infos = doc("body > table:nth-child(3) tr")[2:]
            new_lst = [tuple([td.text_content() for td in tr.cssselect("td")[1:4]]) for tr in infos]
            lst.extend(new_lst)
            logger.debug("第 {} 页记录解析完成, 目前一共有 {} 条信息".format(page, len(lst)))
        logger.info("一共 {} 个同学已提交审核信息".format(len(lst)))
        self.student_list = lst
        return lst

    def export_allocations_info(self, allocs, remove_original=True):
        assert self.s is not None, "登录没有成功"
        for alloc in allocs:
            res = self.s.get(student_url, params={"xh": alloc.student_id})

            # # Optional: 备份
            # if self.backup_dir:
            #     bfile = os.path.join(self.backup_dir, "allocations", "bksjzd_xsgzz_zxj_check_check_xh{}.html".format(alloc.student_id))
            #     with open(bfile, "w") as f:
            #         f.write(res.content)
            #         logger.debug("Write backup html content to {}".format(bfile))

            # Optional: 删除原有助学金
            if remove_original:
                doc = pq(res.content)
                trs = doc("tr")
                text_vs = [tr.text_content() for tr in trs]
                nums = [0, 0, 0]
                parts = [u"院系已批准的校分助学金项目：",
                         u"院系已批准的院分助学金项目：",
                         u"院系已批准的院管助学金项目："]
                type_parts = ["xf", "yf", "yg"]
                for i, part in enumerate(parts):
                    start = text_vs.index(part) + 1
                    while 1 and start < len(trs):
                        lst = trs[start].cssselect("td > font")
                        if not len(lst):
                            break
                        # 助号为第0个
                        delete_url = scho_delete_url.format(student_id=alloc.student_id,
                                                            dm=lst[0].text.strip(),
                                                            type=type_parts[i])
                        res = self.s.get(delete_url)
                        if res.status_code != 200:
                            logger.error("status code {}: 删除分配失败: 学号: {}; 分配: {};\n\t返回: {};"
                                         .format(res.status_code, alloc.student_id,
                                                 " ".join([cell.text.strip() for cell in lst]), res.content))
                        nums[i] += 1
                        start += 1
                logger.debug(u"删除 {}({}) 已有分配:\n\t".format(alloc.name, alloc.student_id) +
                             "\n\t".join([u"删除 {} 条 {}".format(n, p) for n, p in zip(nums, parts)]))

            # Prepare the form
            form = {
                "xh": alloc.student_id,
                "bksjzdaction": "add"
            }
            xf = 0
            yf = 0
            yg = 0
            for sch in alloc.scholars:
                form.update({
                    "dm_add": sch.id,
                })
                if sch.type == u"校管校分":
                    url = xf_url
                    xf += 1
                elif sch.type == u"校管院分":
                    url = yf_url
                    yf += 1
                elif sch.type == u"院管院分":
                    url = yg_url
                    yg += 1

                # Do the request
                res = self.s.post(url, data=form)
                if res.status_code != 200:
                    logger.error("status code {}: 更新分配失败: 学号: {}; 分配: {};\n\t返回: {};"
                                 .format(res.status_code, alloc.student_id, alloc, res.content))
            logger.info(u"Update allocation info for {}({}). 新增个数:\n\t校管校分: {}; 校管院分: {}; 院管院分: {}"
                        .format(alloc.name, alloc.student_id, xf, yf, yg))

AllocationRecord = namedtuple("AllocationRecord", ("name", "cls", "student_id", "scholars"))
Scholarship = namedtuple("Scholarship", ("type", "id"))
legal_scho_types = {u"校管校分", u"校管院分", u"院管院分"}
def parse_allocations(wb):
    allocs_ws = wb.get_sheet_by_name(u"助学金分配名单")
    rows = list(allocs_ws.rows)
    # 年级, 班级, 学号, 姓名, <荣誉列表...>

    records = []
    for row in rows[1:]:
        cell_vs = [cell.value for cell in row]
        scho_cell_vs = cell_vs[4:]
        assert len(scho_cell_vs) % 3 == 0
        start = 0
        scholar_lst = []
        while start < len(scho_cell_vs):
            if not scho_cell_vs[start] or not scho_cell_vs[start].strip():
                start += 3
                continue
            assert scho_cell_vs[start + 2] in legal_scho_types, "不合法的助学金类型: {}; 合法的助学金类型是: {}".format(scho_cell_vs[start+2], legal_scho_types)
            scholar_lst.append(Scholarship(scho_cell_vs[start+2], scho_cell_vs[start]))
            start += 3
        records.append(AllocationRecord(cell_vs[3], cell_vs[1], cell_vs[2], scholar_lst))
    return records

def dump_allocations(wb, ori_wb, allocations_ids):
    ori_ws = ori_wb.get_sheet_by_name(u"助学金分配名单")
    ws = wb.create_sheet(title=u"助学金分配名单")
    rows = list(ori_ws.rows)
    header = [cell.value for cell in rows[0] if cell.value is not None]
    for col, head in enumerate(header):
        ws.cell(row=1, column=col+1).value = head
    row_id = 2
    for ind in allocations_ids:
        cell_vs = [cell.value for cell in rows[ind + 1]]
        for col, v in enumerate(cell_vs):
            ws.cell(row=row_id, column=col+1).value = v
        row_id += 1

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("username", help="管理员学号.")
    parser.add_argument("excel_path", help="Excel路径.")
    parser.add_argument("-v", "--verbose", default=False, action="store_true",
                        help="开启调试信息.")
    parser.add_argument("--verbose-requests", default=False, action="store_true",
                        help="开启http请求调试信息.")
    # parser.add_argument("--backup", default=None,
    #                     help="如果声明了目录, 备份更新前拿到的html结果到某个目录下. 一旦出问题可以让程序员来复原...")
    parser.add_argument("--not-remove-original", default=False, action="store_true",
                        help="如果指定这个选项, 将不会把原来已分配的奖学金删除, 只增加奖学金.")
    parser.add_argument("--not-exists-excel", default="./助学金_录入不存在名单.xlsx", help="如果有不存在学校系统里的学生, 将信息存至这个excel.")
    args = parser.parse_args()

    level = logging.INFO
    if args.verbose:
        level = logging.DEBUG
    logger.setLevel(level)
    if args.verbose_requests:
        requests_logger = logging.getLogger("requests.packages.urllib3")
        requests_logger.setLevel(args.DEBUG)

    # if not args.backup:
    #     print("你没有指定将之前的记录做备份的目录, 如果程序出bug可能无法恢复之前的状态. 如果要指定, 需要提供`--backup [备份目录名]` 命令行参数")
    #     ans = None
    #     while ans not in {"yes", "no"}:
    #         ans = raw_input("是否继续(yes/no)? ").lower()
    #     if ans == "no":
    #         sys.exit(1)

    wb = load_workbook(args.excel_path)
    # Check format: 需要有一个sheets: 助学金分配名单
    assert u'助学金分配名单' in set(wb.sheetnames), "提供的excel需要包含sheet: 助学金分配名单"

    username = args.username
    password = getpass.getpass("Enter the password of the management user: ")
    print("正在登录...")
    # exporter = Exporter(username, password, args.backup)
    exporter = Exporter(username, password, False)

    print("正在获取info上已存在的学生信息... 可能需要较长时间...")
    student_lst = exporter.get_student_list()
    exist_students = [(l[0], l[1]) for l in student_lst] # 学号,姓名,班级
    not_exists_wb = None
    print("正在从excel中解析分配信息...")
    allocations = parse_allocations(wb)
    print("解析完成, 共有 {} 条分配信息.".format(len(allocations)))
    export_students = set([(str(r.student_id), r.name) for r in allocations])
    nonexist_students = export_students - set(exist_students)
    if len(nonexist_students) > 0:
        print(u"以下学生没有在info上存在表项, 暂时导出分配信息时将忽略:\n\t{}".format("\n\t".join([u"{} {}".format(sid, sname) for sid, sname in nonexist_students])))
        export_allocations = []
        notexport_allocations_ids = []
        for i, alloc in enumerate(allocations):
            if (str(alloc.student_id), alloc.name) not in nonexist_students:
                export_allocations.append(alloc)
            else:
                notexport_allocations_ids.append(i)
        not_exists_wb = not_exists_wb or Workbook()
        dump_allocations(not_exists_wb, wb, notexport_allocations_ids)
        print("这些学生的分配信息将被存至 {}".format(args.not_exists_excel))
    else:
        export_allocations = allocations
    print("共有 {} 条分配信息需导入.".format(len(export_allocations)))
    ans = None
    while ans not in {"yes", "no"}:
        ans = raw_input("是否继续(yes/no)? ").lower()
    if ans == "yes":
        print("开始导入到学校系统...")
        exporter.export_allocations_info(export_allocations, (not args.not_remove_original))
    if not_exists_wb is not None:
        not_exists_wb.save(filename=args.not_exists_excel)
